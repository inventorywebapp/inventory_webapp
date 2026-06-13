from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, User, SKU, CountingSession, CountRecord, AuditLog, get_ph_time
from utils import check_recount_needed, sync_data_from_drive
from config import Config
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
import os
import sys
import re
from sqlalchemy import or_

# Configuration for count expiration (in days)
COUNT_EXPIRATION_DAYS = 14

app = Flask(__name__)
app.config.from_object(Config)

# Initialize extensions
db.init_app(app)

# ============================================
# ADD THIS FUNCTION HERE
# ============================================
def ensure_database_schema():
    """Auto-create missing columns/tables without shell access"""
    try:
        import sqlite3
        db_path = app.config.get('SQLALCHEMY_DATABASE_URI', '').replace('sqlite:///', '')
        
        # For in-memory or relative path
        if db_path == '' or db_path == ':memory:':
            db_path = 'instance/inventory.db'
        
        # Create instance folder if needed
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        
        # Connect and add missing schema
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Add is_active column to skus if missing
        cursor.execute("PRAGMA table_info(skus)")
        existing_columns = [col[1] for col in cursor.fetchall()]
        
        if 'is_active' not in existing_columns:
            print("🔧 Adding is_active column to skus...", file=sys.stderr)
            cursor.execute("ALTER TABLE skus ADD COLUMN is_active BOOLEAN DEFAULT 1")
            print("✅ is_active column added", file=sys.stderr)
        
        # Create merge history table if missing
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='sku_merge_history'")
        if not cursor.fetchone():
            print("🔧 Creating sku_merge_history table...", file=sys.stderr)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS sku_merge_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    old_sku_id INTEGER REFERENCES skus(id),
                    new_sku_id INTEGER REFERENCES skus(id),
                    merged_by INTEGER REFERENCES user(id),
                    merged_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    reason VARCHAR(500),
                    records_transferred INTEGER DEFAULT 0
                )
            """)
            print("✅ sku_merge_history table created", file=sys.stderr)
        
        conn.commit()
        conn.close()
        
    except Exception as e:
        print(f"⚠️ Schema check warning: {e}", file=sys.stderr)

# CALL THE FUNCTION HERE (ONCE!)
ensure_database_schema()

# Initialize migrate
migrate = Migrate(app, db)

# Setup login manager
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.session_protection = "strong"

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Create default users
def create_default_users():
    with app.app_context():
        db.create_all()
        
        if User.query.count() == 0:
            admin = User(
                username='admin',
                password=generate_password_hash('admin123'),
                full_name='System Administrator',
                role='admin',
                is_active=True
            )
            db.session.add(admin)
            
            staff = User(
                username='staff1',
                password=generate_password_hash('staff123'),
                full_name='Inventory Staff',
                role='inventory_staff',
                is_active=True
            )
            db.session.add(staff)
            
            supervisor = User(
                username='supervisor1',
                password=generate_password_hash('super123'),
                full_name='Inventory Supervisor',
                role='inventory_supervisor',
                is_active=True
            )
            db.session.add(supervisor)
            
            auditor = User(
                username='auditor1',
                password=generate_password_hash('audit123'),
                full_name='Audit User',
                role='auditor',
                is_active=True
            )
            db.session.add(auditor)
            
            manager = User(
                username='manager1',
                password=generate_password_hash('manager123'),
                full_name='Manager',
                role='manager',
                is_active=True
            )
            db.session.add(manager)
            
            db.session.commit()
            print("Default users created!", file=sys.stderr)
        else:
            print(f"Database already has {User.query.count()} users", file=sys.stderr)

def auto_sync_if_empty():
    """Automatically sync if database is empty"""
    with app.app_context():
        if SKU.query.count() == 0:
            print("Database empty, auto-syncing from Google Drive...", file=sys.stderr)
            try:
                success = sync_data_from_drive()
                if success:
                    print("Auto-sync completed successfully!", file=sys.stderr)
                else:
                    print("Auto-sync failed. Please sync manually.", file=sys.stderr)
            except Exception as e:
                print(f"Auto-sync error: {e}", file=sys.stderr)

# Call the functions
create_default_users()
auto_sync_if_empty()

@app.teardown_appcontext
def shutdown_session(exception=None):
    """Close database session after each request"""
    db.session.remove()

# ... REST OF YOUR CODE CONTINUES (all your routes) ...

@app.route('/')
@login_required
def index():
    if current_user.role == 'admin':
        return redirect(url_for('unified_panel'))
    elif current_user.role == 'manager':
        return redirect(url_for('unified_panel'))
    elif current_user.role == 'inventory_supervisor':
        return redirect(url_for('unified_panel'))
    elif current_user.role == 'auditor':
        return redirect(url_for('unified_panel'))
    elif current_user.has_permission('count'):
        return redirect(url_for('counting'))
    else:
        return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            if not user.is_active:
                flash('Your account is disabled. Please contact administrator.', 'error')
                return redirect(url_for('login'))
            
            login_user(user, remember=True)
            
            audit = AuditLog(
                user_id=user.id,
                action='Login',
                details=f'User {user.full_name} logged in',
                ip_address=request.remote_addr
            )
            db.session.add(audit)
            db.session.commit()
            
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    audit = AuditLog(
        user_id=current_user.id,
        action='Logout',
        details=f'User {current_user.full_name} logged out',
        ip_address=request.remote_addr
    )
    db.session.add(audit)
    db.session.commit()
    
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    total_skus = SKU.query.count()
    active_sessions_count = CountingSession.query.filter_by(is_completed=False).count()
    
    active_session_info = None
    if active_sessions_count > 0:
        latest_session = CountingSession.query.filter_by(is_completed=False).order_by(CountingSession.session_date.desc()).first()
        if latest_session:
            active_session_info = {
                'id': latest_session.id,
                'warehouse': latest_session.warehouse,
                'session_date': latest_session.session_date.strftime('%Y-%m-%d %H:%M:%S'),
                'started_by': latest_session.user.full_name if latest_session.user else 'Unknown'
            }
    
    day_categories_raw = db.session.query(SKU.category).distinct().all()
    day_order = {'Mon': 1, 'Tue': 2, 'Wed': 3, 'Thu': 4, 'Fri': 5, 'Sat': 6}
    day_categories = [c[0] for c in day_categories_raw if c[0] and c[0] != '']
    day_categories.sort(key=lambda x: day_order.get(x, 999))
    
    item_categories_raw = db.session.query(SKU.description).distinct().all()
    item_categories = [c[0] for c in item_categories_raw if c[0] and c[0] != '']
    item_categories.sort()
    
    # REMOVED THE LIMIT - NOW SHOWS ALL ITEM CATEGORIES
    # item_categories = item_categories[:10]  <-- DELETED THIS LINE
    
    warehouses = ['Main Warehouse', '5th Floor Warehouse']
    
    completed_records = db.session.query(CountRecord.sku_id).join(
        CountingSession
    ).filter(
        CountingSession.is_completed == True
    ).distinct().all()
    
    completed_sku_ids = set([r[0] for r in completed_records])
    
    day_category_progress = []
    for day_cat in day_categories:
        sku_ids_in_category = db.session.query(SKU.id).filter(SKU.category == day_cat).all()
        sku_ids = [s[0] for s in sku_ids_in_category]
        total_in_category = len(sku_ids)
        
        completed_count = len([sid for sid in sku_ids if sid in completed_sku_ids])
        
        if total_in_category > 0:
            percentage = (completed_count / total_in_category) * 100
            percentage = round(percentage, 2)
        else:
            percentage = 0
        
        day_category_progress.append({
            'name': day_cat,
            'total_skus': total_in_category,
            'completed_skus': completed_count,
            'percentage': percentage,
            'start_date': 'Not started',
            'end_date': 'In progress'
        })
    
    return render_template('dashboard.html',
                         total_skus=total_skus,
                         active_sessions_count=active_sessions_count,
                         active_session_info=active_session_info,
                         day_categories=day_categories,
                         item_categories=item_categories,  # Now passing ALL categories
                         warehouses=warehouses,
                         day_category_progress=day_category_progress)

@app.route('/get_skus')
@login_required
def get_skus():
    day_category = request.args.get('day_category')
    item_category = request.args.get('item_category')
    search = request.args.get('search', '')
    
    query = SKU.query.filter_by(is_active=True)
    
    is_filtered = False
    
    if day_category and day_category != 'All' and day_category != '-- All Day Categories --':
        query = query.filter(SKU.category == day_category)
        is_filtered = True
    
    if item_category and item_category != 'All' and item_category != '-- All Item Categories --':
        query = query.filter(SKU.description == item_category)
        is_filtered = True
    
    if search and search.strip():
        query = query.filter(
            or_(
                SKU.sku.contains(search), 
                SKU.description.contains(search)
            )
        )
        is_filtered = True
    
    if is_filtered:
        skus = query.all()
    else:
        skus = query.limit(1000).all()
    
    # Calculate expiration cutoff date (timezone-naive for comparison)
    now = get_ph_time()
    expiration_cutoff = now - timedelta(days=COUNT_EXPIRATION_DAYS)
    # Make it timezone-naive for comparison with database times
    expiration_cutoff_naive = expiration_cutoff.replace(tzinfo=None)
    
    result = []
    for sku in skus:
        latest_record = CountRecord.query.filter_by(
            sku_id=sku.id
        ).order_by(CountRecord.version.desc()).first()
        
        final_count = None
        count_time = None
        has_count = False
        is_completed = False
        is_expired = False
        
        if latest_record:
            has_count = True
            # Make count_time timezone-naive for comparison
            if latest_record.count_time:
                count_time_naive = latest_record.count_time.replace(tzinfo=None) if hasattr(latest_record.count_time, 'replace') else latest_record.count_time
                count_time = latest_record.count_time.strftime('%Y-%m-%d %H:%M:%S')
            else:
                count_time_naive = None
            
            # Determine the FINAL validated count
            if latest_record.recount_count and latest_record.recount_count > 0:
                final_count = latest_record.recount_count
            elif latest_record.final_count and latest_record.final_count > 0:
                final_count = latest_record.final_count
            else:
                final_count = latest_record.initial_count
            
            # Check if the session is completed
            session = CountingSession.query.get(latest_record.session_id)
            if session and session.is_completed:
                is_completed = True
            
            # Check if the count has expired (> 14 days old)
            if count_time_naive and count_time_naive < expiration_cutoff_naive:
                is_expired = True
                has_count = False  # Treat as not counted for UI purposes
                is_completed = False
        
        result.append({
            'id': sku.id, 
            'sku': sku.sku, 
            'description': sku.description,
            'category': sku.category, 
            'last_count_date': sku.last_count_date,
            'last_count': sku.last_count, 
            'total_container_qty': sku.total_container_qty,
            'container_details': sku.container_details, 
            'final_expected_count': sku.final_expected_count,
            'kenneth_inventory': sku.kenneth_inventory, 
            'stock_status': sku.stock_status,
            'bypass_recount': sku.bypass_recount,
            'has_count': has_count,
            'final_count': final_count if not is_expired else None,
            'count_time': count_time if not is_expired else None,
            'is_completed': is_completed and not is_expired,
            'is_expired': is_expired
        })
    
    return jsonify(result)

@app.route('/search_suggestions')
@login_required
def search_suggestions():
    search_term = request.args.get('term', '')
    if len(search_term) < 2:
        return jsonify([])
    
    skus = SKU.query.filter_by(is_active=True).filter(
        or_(
            SKU.sku.contains(search_term), 
            SKU.description.contains(search_term)
        )
    ).limit(10).all()
    
    return jsonify([{'sku': s.sku, 'description': s.description, 'id': s.id} for s in skus])

@app.route('/counting', methods=['GET', 'POST'])
@login_required
def counting():
    if not current_user.has_permission('count'):
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        data = request.json
        session_id = data.get('session_id')
        counts = data.get('counts', {})
        warehouse = data.get('warehouse')
        
        session_obj = CountingSession.query.get(session_id)
        if not session_obj:
            session_obj = CountingSession(user_id=current_user.id, warehouse=warehouse)
            db.session.add(session_obj)
            db.session.commit()
            session_id = session_obj.id
        
        recount_needed_skus = []
        for sku_id, count_data in counts.items():
            sku = SKU.query.get(int(sku_id))
            if not sku:
                continue
                
            initial_count = float(count_data.get('initial_count', 0))
            
            pending_recount = CountRecord.query.filter_by(
                session_id=session_obj.id,
                sku_id=int(sku_id),
                recount_completed=False
            ).first()
            
            if pending_recount:
                old_value = pending_recount.initial_count
                old_version = pending_recount.version
                
                pending_recount.initial_count = initial_count
                pending_recount.count_time = get_ph_time()
                
                audit = AuditLog(
                    user_id=current_user.id,
                    action='Count Updated',
                    details=f'SKU {sku.sku}: updated pending recount from {old_value} to {initial_count} (version {old_version})',
                    ip_address=request.remote_addr
                )
                db.session.add(audit)
                
                recount_needed = check_recount_needed(initial_count, sku.final_expected_count, sku.kenneth_inventory)
                pending_recount.is_recount_needed = recount_needed
                
                if recount_needed:
                    recount_needed_skus.append(sku.sku)
            else:
                latest_record = CountRecord.query.filter_by(
                    session_id=session_obj.id, 
                    sku_id=int(sku_id)
                ).order_by(CountRecord.version.desc()).first()
                
                old_value = latest_record.initial_count if latest_record else None
                old_version = latest_record.version if latest_record else 0
                new_version = old_version + 1
                
                recount_needed = check_recount_needed(initial_count, sku.final_expected_count, sku.kenneth_inventory)
                
                new_record = CountRecord(
                    session_id=session_obj.id,
                    sku_id=int(sku_id),
                    initial_count=initial_count,
                    is_recount_needed=recount_needed,
                    count_time=get_ph_time(),
                    version=new_version
                )
                db.session.add(new_record)
                
                if recount_needed:
                    recount_needed_skus.append(sku.sku)
                
                if old_value is not None and old_value != initial_count:
                    audit = AuditLog(
                        user_id=current_user.id,
                        action='Count Changed',
                        details=f'SKU {sku.sku}: version {old_version} count {old_value} → version {new_version} count {initial_count}',
                        ip_address=request.remote_addr
                    )
                    db.session.add(audit)
                else:
                    audit = AuditLog(
                        user_id=current_user.id,
                        action='Count Added',
                        details=f'SKU {sku.sku}: version {new_version} count {initial_count}',
                        ip_address=request.remote_addr
                    )
                    db.session.add(audit)
        
        db.session.commit()
        
        audit = AuditLog(
            user_id=current_user.id,
            action='Initial Count Saved',
            details=f'Saved counts for {len(counts)} SKUs in session {session_obj.id}. Recount needed for {len(recount_needed_skus)} SKUs',
            ip_address=request.remote_addr
        )
        db.session.add(audit)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'session_id': session_obj.id,
            'recount_needed_count': len(recount_needed_skus),
            'recount_needed_skus': recount_needed_skus[:10]
        })
    
    warehouses = ['Main Warehouse', '5th Floor Warehouse']
    day_categories_raw = db.session.query(SKU.category).distinct().all()
    item_categories_raw = db.session.query(SKU.description).distinct().all()
    
    day_order = {'Mon': 1, 'Tue': 2, 'Wed': 3, 'Thu': 4, 'Fri': 5, 'Sat': 6}
    day_categories = [c[0] for c in day_categories_raw if c[0] and c[0] != '']
    day_categories.sort(key=lambda x: day_order.get(x, 999))
    
    item_categories = [c[0] for c in item_categories_raw if c[0] and c[0] != '']
    item_categories.sort()

    # Only show active SKUs in counting page
    active_skus_count = SKU.query.filter_by(is_active=True).count()
    print(f"Active SKUs available for counting: {active_skus_count}", file=sys.stderr)
    
    return render_template('counting.html',
                         warehouses=warehouses,
                         day_categories=day_categories,
                         item_categories=item_categories)

@app.route('/get_recount_list')
@login_required
def get_recount_list():
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify([])
    
    records = CountRecord.query.filter_by(
        session_id=session_id, 
        is_recount_needed=True, 
        recount_completed=False
    ).all()
    
    result = [{
        'id': r.id, 
        'sku_id': r.sku.id, 
        'sku': r.sku.sku, 
        'description': r.sku.description,
        'initial_count': r.initial_count, 
        'final_expected_count': r.sku.final_expected_count,
        'kenneth_inventory': r.sku.kenneth_inventory, 
        'remarks': r.remarks
    } for r in records]
    
    return jsonify(result)

@app.route('/save_recount', methods=['POST'])
@login_required
def save_recount():
    if not current_user.has_permission('recount'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    data = request.json
    for recount_data in data.get('recounts', []):
        record = CountRecord.query.get(recount_data.get('record_id'))
        if record:
            record.recount_count = float(recount_data.get('recount_count', 0))
            record.final_count = record.recount_count
            record.remarks = recount_data.get('remarks', '')
            record.recount_completed = True
            record.is_recount_needed = False
            
            audit = AuditLog(
                user_id=current_user.id,
                action='Recount Completed',
                details=f'SKU {record.sku.sku}: recount count = {record.recount_count}, reason: {record.remarks}',
                ip_address=request.remote_addr
            )
            db.session.add(audit)
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/check_recount_status', methods=['GET'])
@login_required
def check_recount_status():
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({'has_pending_recounts': False})
    
    pending_count = CountRecord.query.filter_by(
        session_id=session_id,
        is_recount_needed=True,
        recount_completed=False
    ).count()
    
    return jsonify({'has_pending_recounts': pending_count > 0, 'count': pending_count})

@app.route('/get_latest_counts', methods=['GET'])
@login_required
def get_latest_counts():
    sku_ids = request.args.get('sku_ids', '')
    session_id = request.args.get('session_id')
    
    if not session_id or not sku_ids:
        return jsonify({})
    
    sku_id_list = sku_ids.split(',') if isinstance(sku_ids, str) else [sku_ids] if sku_ids else []
    
    result = {}
    for sku_id in sku_id_list:
        if not sku_id:
            continue
        latest = CountRecord.query.filter_by(
            session_id=session_id,
            sku_id=int(sku_id)
        ).order_by(CountRecord.version.desc()).first()
        
        if latest:
            result[sku_id] = {
                'initial_count': latest.initial_count,
                'version': latest.version,
                'count_time': latest.count_time.strftime('%Y-%m-%d %H:%M:%S')
            }
    
    return jsonify(result)

@app.route('/complete_counting', methods=['POST'])
@login_required
def complete_counting():
    session_id = request.json.get('session_id')
    if not session_id:
        return jsonify({'success': False, 'message': 'No session ID provided'}), 400
    
    pending_recounts = CountRecord.query.filter_by(
        session_id=session_id,
        is_recount_needed=True,
        recount_completed=False
    ).count()
    
    if pending_recounts > 0:
        return jsonify({
            'success': False, 
            'message': f'Cannot complete session. You have {pending_recounts} pending recount(s) that need to be completed first.'
        }), 400
    
    session_obj = CountingSession.query.get(session_id)
    if session_obj:
        session_obj.is_completed = True
        db.session.commit()
        
        audit = AuditLog(
            user_id=current_user.id,
            action='Complete Counting',
            details=f'Completed counting session {session_id}',
            ip_address=request.remote_addr
        )
        db.session.add(audit)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Counting session completed successfully!'})
    return jsonify({'success': False, 'message': 'Session not found'}), 404

@app.route('/unified_panel')
@login_required
def unified_panel():
    """Unified panel with role-based tab visibility"""
    users = User.query.all() if current_user.role == 'admin' else []
    sessions = CountingSession.query.order_by(CountingSession.session_date.desc()).limit(50).all()
    audit_logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(100).all()
    
    warehouses = ['Main Warehouse', '5th Floor Warehouse', 'All']
    day_categories_raw = db.session.query(SKU.category).distinct().all()
    item_categories_raw = db.session.query(SKU.description).distinct().all()
    
    day_order = {'Mon': 1, 'Tue': 2, 'Wed': 3, 'Thu': 4, 'Fri': 5, 'Sat': 6}
    day_categories = [c[0] for c in day_categories_raw if c[0] and c[0] != '']
    day_categories.sort(key=lambda x: day_order.get(x, 999))
    
    item_categories = [c[0] for c in item_categories_raw if c[0] and c[0] != '']
    item_categories.sort()
    
    return render_template('unified_panel.html', 
                         users=users, 
                         sessions=sessions, 
                         audit_logs=audit_logs,
                         warehouses=warehouses,
                         day_categories=day_categories,
                         item_categories=item_categories)

@app.route('/continue_session/<int:session_id>')
@login_required
def continue_session(session_id):
    """Continue an active counting session with previously counted SKUs"""
    session_obj = CountingSession.query.get(session_id)
    if not session_obj:
        flash('Session not found', 'error')
        return redirect(url_for('counting'))
    
    if session_obj.is_completed:
        flash('This session is already completed', 'warning')
        return redirect(url_for('counting'))
    
    return redirect(url_for('counting', session_id=session_obj.id))

@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('Admin access required', 'error')
        return redirect(url_for('dashboard'))
    
    users = User.query.all()
    sessions = CountingSession.query.order_by(CountingSession.session_date.desc()).limit(50).all()
    audit_logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(100).all()
    
    warehouses = ['Main Warehouse', '5th Floor Warehouse', 'All']
    day_categories_raw = db.session.query(SKU.category).distinct().all()
    item_categories_raw = db.session.query(SKU.description).distinct().all()
    
    day_order = {'Mon': 1, 'Tue': 2, 'Wed': 3, 'Thu': 4, 'Fri': 5, 'Sat': 6}
    day_categories = [c[0] for c in day_categories_raw if c[0] and c[0] != '']
    day_categories.sort(key=lambda x: day_order.get(x, 999))
    
    item_categories = [c[0] for c in item_categories_raw if c[0] and c[0] != '']
    item_categories.sort()
    
    return render_template('admin.html', 
                         users=users, 
                         sessions=sessions, 
                         audit_logs=audit_logs,
                         warehouses=warehouses,
                         day_categories=day_categories,
                         item_categories=item_categories)

@app.route('/admin/users')
@login_required
def admin_users():
    if current_user.role != 'admin':
        flash('Admin access required', 'error')
        return redirect(url_for('dashboard'))
    
    users = User.query.all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/create_user', methods=['POST'])
@login_required
def create_user():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    username = request.form.get('username')
    password = request.form.get('password')
    full_name = request.form.get('full_name')
    role = request.form.get('role')
    
    if User.query.filter_by(username=username).first():
        return jsonify({'success': False, 'message': 'Username already exists'}), 400
    
    new_user = User(
        username=username,
        password=generate_password_hash(password),
        full_name=full_name,
        role=role,
        is_active=True
    )
    db.session.add(new_user)
    db.session.commit()
    
    audit = AuditLog(
        user_id=current_user.id,
        action='User Created',
        details=f'Created user {username} with role {role}',
        ip_address=request.remote_addr
    )
    db.session.add(audit)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'User created successfully'})

@app.route('/admin/edit_user', methods=['POST'])
@login_required
def edit_user():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    user_id = request.form.get('user_id')
    full_name = request.form.get('full_name')
    role = request.form.get('role')
    is_active = request.form.get('is_active') == 'true'
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404
    
    old_role = user.role
    user.full_name = full_name
    user.role = role
    user.is_active = is_active
    
    db.session.commit()
    
    audit = AuditLog(
        user_id=current_user.id,
        action='User Updated',
        details=f'Updated user {user.username}: role {old_role} → {role}, active={is_active}',
        ip_address=request.remote_addr
    )
    db.session.add(audit)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'User updated successfully'})

@app.route('/admin/reset_password', methods=['POST'])
@login_required
def reset_password():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    user_id = request.form.get('user_id')
    new_password = request.form.get('new_password')
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404
    
    user.password = generate_password_hash(new_password)
    db.session.commit()
    
    audit = AuditLog(
        user_id=current_user.id,
        action='Password Reset',
        details=f'Reset password for user {user.username}',
        ip_address=request.remote_addr
    )
    db.session.add(audit)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Password reset successfully'})

@app.route('/admin/delete_user', methods=['POST'])
@login_required
def delete_user():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    user_id = request.form.get('user_id')
    
    if int(user_id) == current_user.id:
        return jsonify({'success': False, 'message': 'Cannot delete your own account'}), 400
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    
    audit = AuditLog(
        user_id=current_user.id,
        action='User Deleted',
        details=f'Deleted user {username}',
        ip_address=request.remote_addr
    )
    db.session.add(audit)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'User deleted successfully'})

@app.route('/export_counts', methods=['POST'])
@login_required
def export_counts():
    if not current_user.has_permission('export_counts'):
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        filter_date = request.form.get('filter_date')
        filter_warehouse = request.form.get('filter_warehouse')
        filter_day_category = request.form.get('filter_day_category')
        filter_item_category = request.form.get('filter_item_category')
        
        sku_query = SKU.query
        
        if filter_day_category and filter_day_category != 'All':
            sku_query = sku_query.filter(SKU.category == filter_day_category)
        if filter_item_category and filter_item_category != 'All':
            sku_query = sku_query.filter(SKU.description == filter_item_category)
        
        all_skus_in_category = sku_query.all()
        
        session_query = CountingSession.query
        
        if filter_date and filter_date.strip():
            try:
                target_date = datetime.strptime(filter_date, '%Y-%m-%d')
                start_date = target_date.replace(hour=0, minute=0, second=0)
                end_date = target_date.replace(hour=23, minute=59, second=59)
                session_query = session_query.filter(CountingSession.session_date.between(start_date, end_date))
            except ValueError:
                pass
        
        if filter_warehouse and filter_warehouse != 'All':
            session_query = session_query.filter(CountingSession.warehouse == filter_warehouse)
        
        sessions = session_query.all()
        
        sku_latest = {}
        
        for sku in all_skus_in_category:
            best_record = None
            best_session = None
            
            for session_obj in sessions:
                record = CountRecord.query.filter_by(
                    session_id=session_obj.id,
                    sku_id=sku.id
                ).order_by(CountRecord.version.desc()).first()
                
                if record:
                    if best_session is None or session_obj.session_date > best_session.session_date:
                        best_record = record
                        best_session = session_obj
            
            sku_latest[sku.id] = {
                'sku': sku,
                'record': best_record,
                'session': best_session
            }
        
        sheets_data = {}
        
        for sku_id, data in sku_latest.items():
            sku = data['sku']
            count_record = data['record']
            session_obj = data['session']
            
            day_category = sku.category or 'Uncategorized'
            
            if day_category not in sheets_data:
                sheets_data[day_category] = []
            
            if count_record and session_obj:
                if count_record.recount_count and count_record.recount_count > 0:
                    final_count = count_record.recount_count
                else:
                    final_count = count_record.initial_count
                
                row_data = {
                    'SKU': str(sku.sku),
                    'Description': str(sku.description) if sku.description else '',
                    'Day Category': str(day_category),
                    'Count Status': 'COMPLETED',
                    'Initial Count': count_record.initial_count,
                    'Recount Count': count_record.recount_count if count_record.recount_count else '',
                    'Final Count': final_count,
                    'Remarks': str(count_record.remarks) if count_record.remarks else '',
                    'Date/Time Counted': count_record.count_time.strftime('%Y-%m-%d %H:%M:%S') if count_record.count_time else '',
                    'Counter': session_obj.user.full_name if session_obj.user else 'Unknown',
                    'Warehouse': session_obj.warehouse if session_obj.warehouse else '',
                    'Session Date': session_obj.session_date.strftime('%Y-%m-%d %H:%M:%S') if session_obj.session_date else '',
                    'Last Count Reference': sku.last_count,
                    'Last Count Date': str(sku.last_count_date) if sku.last_count_date else '',
                    'Final Expected Count': sku.final_expected_count,
                    'Kenneth Inventory': sku.kenneth_inventory
                }
            else:
                row_data = {
                    'SKU': str(sku.sku),
                    'Description': str(sku.description) if sku.description else '',
                    'Day Category': str(day_category),
                    'Count Status': '⚠️ NOT COUNTED ⚠️',
                    'Initial Count': 'NOT COUNTED',
                    'Recount Count': 'N/A',
                    'Final Count': 'PENDING',
                    'Remarks': 'SKU was not counted in this session',
                    'Date/Time Counted': 'NOT COUNTED',
                    'Counter': 'N/A',
                    'Warehouse': filter_warehouse if filter_warehouse and filter_warehouse != 'All' else 'All Warehouses',
                    'Session Date': 'N/A',
                    'Last Count Reference': sku.last_count,
                    'Last Count Date': str(sku.last_count_date) if sku.last_count_date else '',
                    'Final Expected Count': sku.final_expected_count,
                    'Kenneth Inventory': sku.kenneth_inventory
                }
            
            sheets_data[day_category].append(row_data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if sheets_data:
                for sheet_name, data in sheets_data.items():
                    safe_sheet_name = str(sheet_name)[:31].replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace(':', '-')
                    df = pd.DataFrame(data)
                    df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                    
                    worksheet = writer.sheets[safe_sheet_name]
                    status_col = None
                    for idx, col in enumerate(df.columns):
                        if col == 'Count Status':
                            status_col = idx + 1
                            break
                    
                    if status_col:
                        from openpyxl.styles import PatternFill
                        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        
                        for row_idx in range(2, len(data) + 2):
                            cell = worksheet.cell(row=row_idx, column=status_col)
                            if cell.value and 'NOT COUNTED' in str(cell.value):
                                cell.fill = yellow_fill
            else:
                df = pd.DataFrame({'Message': ['No SKU data found for the selected filters']})
                df.to_excel(writer, sheet_name='No Data', index=False)
        
        output.seek(0)
        
        filename_parts = ['inventory_export']
        if filter_date:
            filename_parts.append(filter_date)
        if filter_warehouse and filter_warehouse != 'All':
            filename_parts.append(filter_warehouse.replace(' ', '_'))
        
        filename = '_'.join(filename_parts) + f'_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Export error: {e}", file=sys.stderr)
        flash(f'Export failed: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/export_audit_log', methods=['POST'])
@login_required
def export_audit_log():
    if not current_user.has_permission('export_audit'):
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        filter_user = request.form.get('filter_user')
        
        query = AuditLog.query
        
        if start_date and start_date.strip():
            try:
                start = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(AuditLog.timestamp >= start)
            except ValueError:
                pass
        
        if end_date and end_date.strip():
            try:
                end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(AuditLog.timestamp < end)
            except ValueError:
                pass
        
        if filter_user and filter_user != 'All':
            try:
                query = query.filter(AuditLog.user_id == int(filter_user))
            except ValueError:
                pass
        
        logs = query.order_by(AuditLog.timestamp.desc()).all()
        
        if not logs:
            flash('No audit log data found for the selected filters', 'warning')
            return redirect(url_for('admin_dashboard'))
        
        audit_data = []
        for log in logs:
            details = log.details
            if details:
                details = re.sub(r'(\d+)\.0', r'\1', details)
            
            audit_data.append({
                'Timestamp (PHT)': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'User': log.user.full_name if log.user else 'System',
                'Action': log.action,
                'Details': details,
                'IP Address': log.ip_address
            })
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df = pd.DataFrame(audit_data)
            df.to_excel(writer, sheet_name='Audit Log', index=False)
        
        output.seek(0)
        
        filename = f"audit_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Audit export error: {e}", file=sys.stderr)
        flash(f'Audit export failed: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/audit')
@login_required
def audit_dashboard():
    if current_user.role != 'auditor' and current_user.role != 'admin' and not current_user.has_permission('export_audit'):
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    return render_template('audit.html')

@app.route('/get_audit_logs')
@login_required
def get_audit_logs():
    if current_user.role not in ['admin', 'auditor'] and not current_user.has_permission('export_audit'):
        return jsonify([])
    
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(200).all()
    result = []
    for log in logs:
        details = log.details
        if details:
            details = re.sub(r'(\d+)\.0', r'\1', details)
        
        result.append({
            'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'user': log.user.full_name if log.user else 'Unknown',
            'action': log.action,
            'details': details,
            'ip_address': log.ip_address
        })
    
    return jsonify(result)

@app.route('/api/cleanup', methods=['POST'])
@login_required
def api_cleanup():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    days = int(request.form.get('days', 30))
    cutoff = get_ph_time() - timedelta(days=days)
    
    old_counts = CountRecord.query.filter(CountRecord.count_time < cutoff).delete()
    old_sessions = CountingSession.query.filter(CountingSession.session_date < cutoff).delete()
    old_audits = AuditLog.query.filter(AuditLog.timestamp < cutoff).delete()
    db.session.commit()
    
    return jsonify({'message': f'Deleted {old_counts} counts, {old_sessions} sessions, {old_audits} audit logs', 'success': True})

@app.route('/sync_data', methods=['POST'])
@login_required
def sync_data():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        success = sync_data_from_drive()
        
        if success:
            return jsonify({'success': True, 'message': 'Data synced successfully from Google Drive!'})
        else:
            return jsonify({'success': False, 'message': 'Sync failed. Check server logs for details.'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# Debug endpoints (admin only)
@app.route('/debug_counts/<int:sku_id>')
@login_required
def debug_counts(sku_id):
    if current_user.role != 'admin':
        return "Unauthorized", 403
    
    session_id = request.args.get('session_id')
    if not session_id:
        return "Need session_id parameter"
    
    records = CountRecord.query.filter_by(
        session_id=session_id,
        sku_id=sku_id
    ).order_by(CountRecord.version).all()
    
    sku = SKU.query.get(sku_id)
    
    result = f"<h2>Count History for SKU: {sku.sku if sku else 'Unknown'}</h2>"
    result += "<table border='1' cellpadding='5'>"
    result += "<tr><th>Version</th><th>Count</th><th>Recount</th><th>Final</th><th>Time</th></tr>"
    
    for r in records:
        result += f"<tr>"
        result += f"日上午{r.version}</td>"
        result += f"<td>{r.initial_count}</td>"
        result += f"<td>{r.recount_count if r.recount_count else '-'}</td>"
        result += f"<td>{r.final_count if r.final_count else '-'}</td>"
        result += f"<td>{r.count_time}</td>"
        result += f"</tr>"
    
    result += "<tr>"
    result += f"<p><strong>Latest count: {records[-1].initial_count if records else 'None'}</strong></p>"
    result += '<p><a href="/admin">Back to Admin</a></p>'
    
    return result

@app.route('/debug_session/<int:session_id>')
@login_required
def debug_session(session_id):
    if current_user.role != 'admin':
        return "Unauthorized", 403
    
    records = CountRecord.query.filter_by(session_id=session_id).order_by(CountRecord.sku_id, CountRecord.version).all()
    
    result = f"<h2>Session {session_id} - All Count Records</h2>"
    result += "<table border='1' cellpadding='5'>"
    result += "<tr><th>SKU</th><th>Version</th><th>Count</th><th>Time</th><th>Recount</th><th>Final</th></tr>"
    
    for r in records:
        result += f"<tr>"
        result += f"日上午{r.sku.sku if r.sku else 'Unknown'}</td>"
        result += f"日上午{r.version}</td>"
        result += f"<td>{r.initial_count}</td>"
        result += f"<td>{r.count_time}</td>"
        result += f"<td>{r.recount_count if r.recount_count else '-'}</td>"
        result += f"<td>{r.final_count if r.final_count else '-'}</td>"
        result += f"</tr>"
    
    result += "</table>"
    result += '<p><a href="/admin">Back to Admin</a></p>'
    
    return result

@app.route('/get_in_progress_skus/<int:session_id>')
@login_required
def get_in_progress_skus(session_id):
    """Get SKUs that have been counted but session not completed"""
    session_obj = CountingSession.query.get(session_id)
    if not session_obj:
        return jsonify({'error': 'Session not found'}), 404
    
    counted_sku_ids = db.session.query(CountRecord.sku_id).filter(
        CountRecord.session_id == session_id
    ).distinct().all()
    
    counted_ids = [c[0] for c in counted_sku_ids]
    
    if not counted_ids:
        return jsonify({'skus': [], 'message': 'No SKUs have been counted yet'})
    
    skus_data = []
    for sku_id in counted_ids:
        sku = SKU.query.get(sku_id)
        latest_record = CountRecord.query.filter_by(
            session_id=session_id,
            sku_id=sku_id
        ).order_by(CountRecord.version.desc()).first()
        
        if sku and latest_record:
            # Get the final count (recount if exists)
            if latest_record.recount_count and latest_record.recount_count > 0:
                final_count = latest_record.recount_count
            elif latest_record.final_count and latest_record.final_count > 0:
                final_count = latest_record.final_count
            else:
                final_count = latest_record.initial_count
            
            skus_data.append({
                'id': sku.id,
                'sku': sku.sku,
                'description': sku.description,
                'category': sku.category,
                'last_count_date': sku.last_count_date,
                'last_count': sku.last_count,
                'total_container_qty': sku.total_container_qty,
                'container_details': sku.container_details,
                'final_expected_count': sku.final_expected_count,
                'kenneth_inventory': sku.kenneth_inventory,
                'stock_status': sku.stock_status,
                'bypass_recount': sku.bypass_recount,
                'has_count': True,
                'current_count': latest_record.initial_count,
                'final_count': final_count,
                'count_time': latest_record.count_time.strftime('%Y-%m-%d %H:%M:%S') if latest_record.count_time else '',
                'version': latest_record.version,
                'in_progress': True
            })
    
    return jsonify({'skus': skus_data, 'count': len(skus_data)})

# Add this endpoint after your existing routes and before the health check
@app.route('/api/active_session')
@login_required
def get_active_session():
    """Get the current active session for the user"""
    session = CountingSession.query.filter_by(
        user_id=current_user.id,
        is_completed=False
    ).order_by(CountingSession.session_date.desc()).first()
    
    if session:
        return jsonify({'session_id': session.id})
    return jsonify({'session_id': None})

# Health check endpoint for Render
@app.route('/health')
def health_check():
    return jsonify({'status': 'healthy'}), 200

@app.route('/debug/sync_info')
@login_required
def debug_sync_info():
    """Debug endpoint to check sync status and column mapping"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    from utils import get_sync_status, detect_column_mapping
    import pandas as pd
    import io
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    
    result = {
        'sync_status': get_sync_status(),
        'database_stats': {
            'total_skus': SKU.query.count(),
            'sample_skus': []
        },
        'excel_info': {},
        'column_mapping': {}
    }
    
    # Get sample SKUs
    sample = SKU.query.limit(5).all()
    for sku in sample:
        result['database_stats']['sample_skus'].append({
            'sku': sku.sku,
            'description': sku.description,
            'category': sku.category
        })
    
    # Try to read Excel to check columns
    try:
        creds_json = os.environ.get('GOOGLE_CREDENTIALS_JSON')
        folder_id = os.environ.get('GOOGLE_DRIVE_FOLDER_ID')
        
        if creds_json and folder_id:
            creds_dict = json.loads(creds_json)
            credentials = service_account.Credentials.from_service_account_info(
                creds_dict, 
                scopes=['https://www.googleapis.com/auth/drive.readonly']
            )
            drive_service = build('drive', 'v3', credentials=credentials)
            
            results = drive_service.files().list(
                q=f"'{folder_id}' in parents and (mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or mimeType='application/vnd.ms-excel')",
                fields="files(id, name, createdTime)",
                orderBy="createdTime desc"
            ).execute()
            
            files = results.get('files', [])
            if files:
                latest_file = files[0]
                result['excel_info']['file_name'] = latest_file['name']
                result['excel_info']['file_id'] = latest_file['id']
                result['excel_info']['created'] = latest_file.get('createdTime')
                
                # Download and read first few rows
                request = drive_service.files().get_media(fileId=latest_file['id'])
                file_stream = io.BytesIO()
                downloader = MediaIoBaseDownload(file_stream, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                
                file_stream.seek(0)
                df = pd.read_excel(file_stream)
                
                result['excel_info']['total_rows'] = len(df)
                result['excel_info']['columns'] = list(df.columns)
                result['excel_info']['sample_data'] = df.head(3).to_dict('records')
                
                # Detect column mapping
                result['column_mapping'] = detect_column_mapping(df)
    except Exception as e:
        result['error'] = str(e)
    
    return jsonify(result)

@app.route('/debug/force_refresh')
@login_required
def debug_force_refresh():
    """Force refresh database connection and clear any cache"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        # Force a database commit to ensure all changes are persisted
        db.session.commit()
        
        # Get current count
        sku_count = SKU.query.count()
        
        return jsonify({
            'success': True,
            'message': 'Database refreshed',
            'total_skus': sku_count,
            'timestamp': get_ph_time().strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ============ SKU MERGE TOOL ENDPOINTS ============

@app.route('/admin/merge_skus')
@login_required
def admin_merge_skus():
    """Admin tool to merge/transfer SKU history"""
    if current_user.role != 'admin':
        flash('Admin access required', 'error')
        return redirect(url_for('dashboard'))
    
    from models import SKUMergeHistory
    
    # Get all inactive SKUs (potential old codes)
    inactive_skus = SKU.query.filter_by(is_active=False).order_by(SKU.sku).all()
    
    # Get all active SKUs (potential new codes)
    active_skus = SKU.query.filter_by(is_active=True).order_by(SKU.sku).all()
    
    # Get merge history
    merge_history = SKUMergeHistory.query.order_by(SKUMergeHistory.merged_at.desc()).limit(50).all()
    
    return render_template('admin_merge_skus.html', 
                         inactive_skus=inactive_skus,
                         active_skus=active_skus,
                         merge_history=merge_history)


@app.route('/admin/transfer_history', methods=['POST'])
@login_required
def transfer_history():
    """Transfer count history from old SKU to new SKU"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    from models import SKUMergeHistory
    
    old_sku_id = request.form.get('old_sku_id')
    new_sku_id = request.form.get('new_sku_id')
    reason = request.form.get('reason', 'SKU code renamed/updated')
    
    old_sku = SKU.query.get(old_sku_id)
    new_sku = SKU.query.get(new_sku_id)
    
    if not old_sku or not new_sku:
        return jsonify({'success': False, 'message': 'SKU not found'}), 404
    
    if old_sku.id == new_sku.id:
        return jsonify({'success': False, 'message': 'Cannot transfer to the same SKU'}), 400
    
    try:
        # 1. Transfer all CountRecords from old SKU to new SKU
        count_records = CountRecord.query.filter_by(sku_id=old_sku.id).all()
        transferred_count = 0
        
        for record in count_records:
            # Check if new SKU already has a record for this session/version
            existing = CountRecord.query.filter_by(
                session_id=record.session_id,
                sku_id=new_sku.id,
                version=record.version
            ).first()
            
            if not existing:
                # Create new record for new SKU
                new_record = CountRecord(
                    session_id=record.session_id,
                    sku_id=new_sku.id,
                    initial_count=record.initial_count,
                    recount_count=record.recount_count,
                    final_count=record.final_count,
                    remarks=f"[TRANSFERRED FROM {old_sku.sku}] {record.remarks if record.remarks else ''}",
                    is_recount_needed=record.is_recount_needed,
                    recount_completed=record.recount_completed,
                    count_time=record.count_time,
                    version=record.version
                )
                db.session.add(new_record)
                transferred_count += 1
        
        # 2. Mark old SKU as inactive (if not already)
        old_sku.is_active = False
        
        # 3. Record the merge
        merge_record = SKUMergeHistory(
            old_sku_id=old_sku.id,
            new_sku_id=new_sku.id,
            merged_by=current_user.id,
            reason=reason,
            records_transferred=transferred_count
        )
        db.session.add(merge_record)
        
        db.session.commit()
        
        # 4. Audit log
        audit = AuditLog(
            user_id=current_user.id,
            action='SKU History Transferred',
            details=f"Transferred {transferred_count} count records from {old_sku.sku} to {new_sku.sku}. Reason: {reason}",
            ip_address=request.remote_addr
        )
        db.session.add(audit)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Successfully transferred {transferred_count} count records from {old_sku.sku} to {new_sku.sku}',
            'transferred_count': transferred_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@app.route('/admin/sku_details/<int:sku_id>')
@login_required
def sku_details(sku_id):
    """View detailed history of a SKU"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    sku = SKU.query.get(sku_id)
    if not sku:
        return jsonify({'error': 'SKU not found'}), 404
    
    # Get all count records
    count_records = CountRecord.query.filter_by(sku_id=sku.id).order_by(CountRecord.count_time.desc()).limit(50).all()
    
    records_data = []
    for record in count_records:
        session = CountingSession.query.get(record.session_id)
        records_data.append({
            'session_id': record.session_id,
            'date': record.count_time.strftime('%Y-%m-%d %H:%M:%S') if record.count_time else 'Unknown',
            'initial_count': record.initial_count,
            'recount_count': record.recount_count,
            'final_count': record.final_count,
            'completed': session.is_completed if session else False,
            'counter': session.user.full_name if session and session.user else 'Unknown'
        })
    
    return jsonify({
        'sku': sku.sku,
        'description': sku.description,
        'category': sku.category,
        'is_active': sku.is_active,
        'total_counts': len(count_records),
        'count_records': records_data
    })


@app.route('/admin/suggest_sku_match')
@login_required
def suggest_sku_match():
    """Suggest which active SKU might match an inactive SKU"""
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    inactive_sku_id = request.args.get('sku_id')
    if not inactive_sku_id:
        return jsonify([])
    
    inactive_sku = SKU.query.get(inactive_sku_id)
    if not inactive_sku:
        return jsonify([])
    
    # Find similar active SKUs based on name similarity
    active_skus = SKU.query.filter_by(is_active=True).all()
    
    suggestions = []
    for active in active_skus:
        # Calculate similarity
        similarity = 0
        
        # Check if SKU codes share common parts
        inactive_parts = set(inactive_sku.sku.lower().replace('-', ' ').split())
        active_parts = set(active.sku.lower().replace('-', ' ').split())
        
        common_parts = inactive_parts.intersection(active_parts)
        if common_parts:
            similarity += len(common_parts) * 20
        
        # Check description similarity
        if inactive_sku.description and active.description:
            if inactive_sku.description.lower() == active.description.lower():
                similarity += 50
            elif inactive_sku.description.lower() in active.description.lower() or active.description.lower() in inactive_sku.description.lower():
                similarity += 30
        
        # Check category match
        if inactive_sku.category and active.category and inactive_sku.category == active.category:
            similarity += 15
        
        if similarity > 0:
            suggestions.append({
                'id': active.id,
                'sku': active.sku,
                'description': active.description,
                'category': active.category,
                'similarity': similarity
            })
    
    # Sort by similarity (highest first)
    suggestions.sort(key=lambda x: x['similarity'], reverse=True)
    
    return jsonify(suggestions[:5])


@app.route('/admin/inactive_skus')
@login_required
def admin_inactive_skus():
    """View all inactive SKUs"""
    if current_user.role != 'admin':
        flash('Admin access required', 'error')
        return redirect(url_for('dashboard'))
    
    inactive_skus = SKU.query.filter_by(is_active=False).order_by(SKU.sku).all()
    return render_template('admin_inactive_skus.html', inactive_skus=inactive_skus)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
